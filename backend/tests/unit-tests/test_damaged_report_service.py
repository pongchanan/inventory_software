import pytest
from unittest.mock import MagicMock, patch, call
from datetime import datetime

from app.services.damaged_report_service import (
    get_all_reports,
    get_reports_by_user,
    get_report_image_url,
    create_user_report,
    create_admin_report,
    export_reports_excel,
)


_FAKE_KEY = "damaged-reports/user_1_abcd1234.jpg"
_FAKE_URL = "https://s3.example.com/presigned"


class TestGetAllReports:
    def test_returns_list(self, mock_db, sample_damaged_report):
        mock_db.query.return_value.order_by.return_value.all.return_value = [
            sample_damaged_report
        ]

        result = get_all_reports(mock_db)
        assert result == [sample_damaged_report]

    def test_empty(self, mock_db):
        mock_db.query.return_value.order_by.return_value.all.return_value = []

        result = get_all_reports(mock_db)
        assert result == []


class TestGetReportsByUser:
    def test_filters_by_user(self, mock_db, sample_damaged_report):
        mock_db.query.return_value.filter.return_value.order_by.return_value.all.return_value = [
            sample_damaged_report
        ]

        result = get_reports_by_user(mock_db, user_id=1)
        assert result == [sample_damaged_report]

    def test_empty_for_user(self, mock_db):
        mock_db.query.return_value.filter.return_value.order_by.return_value.all.return_value = (
            []
        )

        result = get_reports_by_user(mock_db, user_id=99)
        assert result == []


class TestGetReportImageUrl:
    @patch(
        "app.services.damaged_report_service.get_presigned_url", return_value=_FAKE_URL
    )
    def test_success(self, mock_presigned, mock_db, sample_damaged_report):
        mock_db.query.return_value.filter.return_value.first.return_value = (
            sample_damaged_report
        )

        url = get_report_image_url(mock_db, 1)

        mock_presigned.assert_called_once_with(sample_damaged_report.illustrated_path)
        assert url == _FAKE_URL

    def test_report_not_found(self, mock_db):
        mock_db.query.return_value.filter.return_value.first.return_value = None

        with pytest.raises(ValueError, match="not found"):
            get_report_image_url(mock_db, 99)


class TestCreateUserReport:
    @patch(
        "app.services.damaged_report_service._upload_damaged_image",
        return_value=_FAKE_KEY,
    )
    def test_success(self, mock_upload, mock_db, sample_borrowing):
        sample_borrowing.item_id = 5
        sample_borrowing.return_at = None
        mock_db.query.return_value.filter.return_value.first.return_value = (
            sample_borrowing
        )

        result = create_user_report(
            mock_db, user_id=1, topic="Cracked", description="Crack", image_data=b"img"
        )

        mock_upload.assert_called_once_with(b"img", 1)
        mock_db.add.assert_called_once()
        mock_db.commit.assert_called_once()
        mock_db.refresh.assert_called_once()
        added = mock_db.add.call_args[0][0]
        assert added.item_id == 5
        assert added.report_by == 1
        assert added.illustrated_path == _FAKE_KEY

    def test_no_active_borrowing(self, mock_db):
        mock_db.query.return_value.filter.return_value.first.return_value = None

        with pytest.raises(ValueError, match="no active borrowing"):
            create_user_report(
                mock_db, user_id=1, topic="X", description="Y", image_data=b"img"
            )
        mock_db.add.assert_not_called()


class TestCreateAdminReport:
    @patch(
        "app.services.damaged_report_service._upload_damaged_image",
        return_value=_FAKE_KEY,
    )
    def test_success_decrements_quantity(self, mock_upload, mock_db, sample_item):
        sample_item.quantity = 3
        mock_db.query.return_value.filter.return_value.first.return_value = sample_item

        create_admin_report(
            mock_db,
            admin_id=99,
            item_id=1,
            topic="Dent",
            description="Big dent",
            image_data=b"img",
        )

        assert sample_item.quantity == 2
        mock_db.add.assert_called_once()
        mock_db.commit.assert_called_once()
        mock_db.refresh.assert_called_once()
        added = mock_db.add.call_args[0][0]
        assert added.item_id == 1
        assert added.report_by == 99
        assert added.illustrated_path == _FAKE_KEY

    @patch(
        "app.services.damaged_report_service._upload_damaged_image",
        return_value=_FAKE_KEY,
    )
    def test_quantity_floors_at_zero(self, mock_upload, mock_db, sample_item):
        sample_item.quantity = 0
        mock_db.query.return_value.filter.return_value.first.return_value = sample_item

        create_admin_report(
            mock_db,
            admin_id=99,
            item_id=1,
            topic="X",
            description="Y",
            image_data=b"img",
        )

        assert sample_item.quantity == 0

    def test_item_not_found(self, mock_db):
        mock_db.query.return_value.filter.return_value.first.return_value = None

        with pytest.raises(ValueError, match="not found"):
            create_admin_report(
                mock_db,
                admin_id=99,
                item_id=999,
                topic="X",
                description="Y",
                image_data=b"img",
            )
        mock_db.add.assert_not_called()


class TestExportReportsExcel:
    def _make_row(
        self,
        report_id=1,
        topic="Cracked",
        desc="visible crack",
        item_name="Drill",
        reporter="Alice",
        key="damaged-reports/x.jpg",
    ):
        report = MagicMock()
        report.id = report_id
        report.topic = topic
        report.description = desc
        report.report_at = datetime(2026, 4, 1, 9, 0)
        report.illustrated_path = key
        return (report, item_name, reporter)

    def _setup_query(self, mock_db, rows):
        (
            mock_db.query.return_value.join.return_value.join.return_value.order_by.return_value.all.return_value
        ) = rows

    def test_returns_bytes(self, mock_db):
        self._setup_query(mock_db, [self._make_row()])

        result = export_reports_excel(mock_db)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_excel_has_headers(self, mock_db):
        import openpyxl, io

        self._setup_query(mock_db, [self._make_row()])

        data = export_reports_excel(mock_db)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, 8)]
        assert headers == [
            "ID",
            "Topic",
            "Description",
            "Item",
            "Reported By",
            "Reported At",
            "Image Key",
        ]

    def test_excel_row_values(self, mock_db):
        import openpyxl, io

        self._setup_query(
            mock_db, [self._make_row(report_id=7, item_name="Drill", reporter="Alice")]
        )

        data = export_reports_excel(mock_db)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.cell(2, 1).value == 7
        assert ws.cell(2, 4).value == "Drill"
        assert ws.cell(2, 5).value == "Alice"

    def test_empty_export(self, mock_db):
        import openpyxl, io

        self._setup_query(mock_db, [])

        data = export_reports_excel(mock_db)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.max_row == 1  # only header row
